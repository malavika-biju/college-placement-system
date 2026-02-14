from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count, Q
from guestapp.models import tbl_company, tbl_login, tbl_student
from .models import tbl_district, tbl_batch, tbl_department, tbl_classtype, tbl_location, tbl_course, tbl_trainingclass, tbl_requests, tbl_student_schedule, tbl_batch, tbl_department, tbl_course
from companyapp.models import tbl_jobpost
from django.core.paginator import Paginator
from datetime import date
from catalystProject.constants import STUDENT_SCHEDULE_STATUS, REQUEST_STATUS, LOGIN_STATUS
from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.styles import Font, PatternFill
import json


def validate_request_creation(jobpost, batch, student_count=0):
    today = date.today()

    if jobpost.application_end_date < today:
        raise ValidationError("Job application deadline has passed")

    if jobpost.company_id.login_id.status != LOGIN_STATUS['CONFIRMED']:
        raise ValidationError("Company is not approved")

    eligible_students = tbl_student.objects.filter(
        login_id__status='confirmed',
        percentage__gte=jobpost.cutoff_mark,
        batch_id=batch
    )

    eligible_count = eligible_students.count()

    if student_count and student_count > eligible_count:
        raise ValidationError(
            f"Requested {student_count} students but only {eligible_count} are eligible"
        )

    return eligible_count





def admin_dashboard(request):
    """Clean admin dashboard with placement reports"""
    try:
        # Get dashboard statistics
        total_students = tbl_student.objects.filter(login_id__status='confirmed').count()
        total_companies = tbl_company.objects.filter(login_id__status='confirmed').count()
        total_job_posts = tbl_jobpost.objects.all().count()
        pending_requests = tbl_requests.objects.filter(status='pending').count()
        
        # Get placement data for charts
        courses = tbl_course.objects.all().order_by('course_name')
        chart_data = []
        
        for course in courses:
            # Count placed students for this course
            placed_count = tbl_student_schedule.objects.filter(
                student_id__course_id=course,
                status='selected'
            ).count()
            
            # Count total students in this course
            total_course_students = tbl_student.objects.filter(
                course_id=course,
                login_id__status='confirmed'
            ).count()
            
            # Calculate placement percentage
            placement_percent = 0
            if total_course_students > 0:
                placement_percent = round((placed_count / total_course_students) * 100, 2)
            
            chart_data.append({
                'course_name': course.course_name,
                'placed': placed_count,
                'total': total_course_students,
                'percentage': placement_percent,
                'department': course.department_id.department_name if course.department_id else "Computer Science"
            })
        
        # Sort by placed count (highest first)
        chart_data.sort(key=lambda x: x['placed'], reverse=True)
        
        # Calculate totals
        total_placed = sum(item['placed'] for item in chart_data)
        total_students_all = sum(item['total'] for item in chart_data)
        
        # Calculate overall placement percentage
        overall_percentage = 0
        if total_students_all > 0:
            overall_percentage = round((total_placed / total_students_all) * 100, 2)
        
        # Prepare data for Chart.js
        course_names = [item['course_name'] for item in chart_data]
        placed_counts = [item['placed'] for item in chart_data]
        percentages = [item['percentage'] for item in chart_data]
        
        # Get student distribution data for pie chart
        student_distribution = []
        for course in courses:
            student_count = tbl_student.objects.filter(
                course_id=course,
                login_id__status='confirmed'
            ).count()
            if student_count > 0:
                student_distribution.append({
                    'course_name': course.course_name,
                    'count': student_count,
                    'department': course.department_id.department_name if course.department_id else "General"
                })
        
        # Sort by student count (descending)
        student_distribution.sort(key=lambda x: x['count'], reverse=True)
        
        # Prepare data for pie chart
        distribution_labels = [item['course_name'] for item in student_distribution]
        distribution_data = [item['count'] for item in student_distribution]
        
        # Get recent activities
        recent_companies = tbl_company.objects.filter(
            login_id__status='confirmed'
        ).select_related('location_id').order_by('-reg_date')[:5]
        
        recent_jobs = tbl_jobpost.objects.order_by('-post_date')[:5]
        
        context = {
            # Statistics cards
            'total_students': total_students,
            'total_companies': total_companies,
            'total_job_posts': total_job_posts,
            'pending_requests': pending_requests,
            
            # Chart data
            'chart_data': chart_data,
            'course_names_json': json.dumps(course_names),
            'placed_counts_json': json.dumps(placed_counts),
            'percentages_json': json.dumps(percentages),  # Fixed typo: was 'percentiles_json'
            
            # Student distribution for pie chart
            'student_distribution': student_distribution,
            'distribution_labels_json': json.dumps(distribution_labels),
            'distribution_data_json': json.dumps(distribution_data),
            
            # Totals and percentages
            'total_placed': total_placed,
            'total_students_all': total_students_all,
            'total_courses': len(chart_data),
            'overall_percentage': overall_percentage,
            'placement_rate': overall_percentage,  # Same as overall_percentage
            
            # Recent activities
            'recent_companies': recent_companies,
            'recent_jobs': recent_jobs,
        }
        
        return render(request, 'Admin/dashboard.html', context)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f"""
            <script>
                alert('Error loading dashboard: {str(e)}');
                window.location='/adminapp/admin_dashboard/';
            </script>
        """)
def district(request):
    return render(request, "Admin/district.html")


def district_insert(request):
    if request.method == "POST":
        dname = request.POST.get("district_name")

        if tbl_district.objects.filter(district_name=dname).exists():
            return HttpResponse("<script>alert('District already exists');window.location='/Admin/district';</script>")

        obj = tbl_district(district_name=dname)
        obj.save()

        return HttpResponse("<script>alert('District added successfully');window.location='/adminapp/district';</script>")

    return HttpResponse("<script>alert('Invalid Request');window.location='/adminapp/district';</script>")
def view_districts(request):
    data = tbl_district.objects.all().order_by('district_name')
    return render(request, "Admin/view_districts.html", {"districts": data})

def delete_district(request, district_id):
    try:
        district = tbl_district.objects.get(district_id=district_id)
        district.delete()
        return HttpResponse("<script>alert('District deleted successfully');window.location='/adminapp/district/view/';</script>")
    except tbl_district.DoesNotExist:
        return HttpResponse("<script>alert('District not found');window.location='/adminapp/district/view/';</script>")
    
def edit_district(request, district_id):
    try:
        district = tbl_district.objects.get(district_id=district_id)
    except tbl_district.DoesNotExist:
        return HttpResponse("<script>alert('District not found');window.location='/adminapp/district/view/';</script>")

    if request.method == "POST":
        new_name = request.POST.get("district_name")

        if tbl_district.objects.filter(district_name=new_name).exclude(district_id=district_id).exists():
            return HttpResponse("<script>alert('District name already exists');window.location='/adminapp/edit_district/{}';</script>".format(district_id))

        district.district_name = new_name
        district.save()
        return HttpResponse("<script>alert('District updated successfully');window.location='/adminapp/district/view/';</script>")

    return render(request, "Admin/edit_district.html", {"district": district})

def batch(request):
    return render(request, "Admin/batch.html")


def batch_insert(request):
    if request.method == "POST":
        byear = request.POST.get("batch_year")

        if tbl_batch.objects.filter(batch_year=byear).exists():
            return HttpResponse("<script>alert('Batch already exists');window.location='/adminapp/batch/';</script>")

        obj = tbl_batch(batch_year=byear)
        obj.save()

        return HttpResponse("<script>alert('Batch added successfully');window.location='/adminapp/batch/';</script>")

    return HttpResponse("<script>alert('Invalid Request');window.location='/adminapp/batch/';</script>")

def view_batch(request):
    data = tbl_batch.objects.all().order_by('batch_year')
    return render(request, "Admin/view_batch.html", {"batches": data})
def delete_batch(request, batch_id):
    try:
        batch = tbl_batch.objects.get(batch_id=batch_id)
        batch.delete()
        return HttpResponse("<script>alert('Batch deleted successfully');window.location='/adminapp/view_batch/';</script>")
    except tbl_batch.DoesNotExist:
        return HttpResponse("<script>alert('Batch not found');window.location='/adminapp/view_batch/';</script>")
    
def edit_batch(request, batch_id):
    try:
        batch = tbl_batch.objects.get(batch_id=batch_id)
    except tbl_batch.DoesNotExist:
        return HttpResponse("<script>alert('Batch not found');window.location='/adminapp/view_batch/';</script>")

    if request.method == "POST":
        new_year = request.POST.get("batch_year")

        if tbl_batch.objects.filter(batch_year=new_year).exclude(batch_id=batch_id).exists():
            return HttpResponse("<script>alert('Batch year already exists');window.location='/adminapp/edit_batch/{}';</script>".format(batch_id))

        batch.batch_year = new_year
        batch.save()
        return HttpResponse("<script>alert('Batch updated successfully');window.location='/adminapp/view_batch/';</script>")

    return render(request, "Admin/edit_batch.html", {"batch": batch})



def department(request):
    return render(request, "Admin/department.html")


def department_insert(request):
    if request.method == "POST":
        name = request.POST.get("department_name")

        if tbl_department.objects.filter(department_name=name).exists():
            return HttpResponse(
                "<script>alert('Department already exists');window.location='/adminapp/department/';</script>"
            )

        tbl_department.objects.create(department_name=name)

        return HttpResponse(
            "<script>alert('Department added successfully');window.location='/adminapp/department/';</script>"
        )

    return HttpResponse(
        "<script>alert('Invalid Request');window.location='/adminapp/department/';</script>"
    )


def view_department(request):
    data = tbl_department.objects.all().order_by("department_name")
    return render(request, "Admin/view_department.html", {"departments": data})

def delete_department(request, department_id):

    try:
        department = tbl_department.objects.get(department_id=department_id)
        department.delete()
        return HttpResponse("<script>alert('Department deleted successfully');window.location='/adminapp/view_department/';</script>")
    except tbl_department.DoesNotExist:
        return HttpResponse("<script>alert('Department not found');window.location='/adminapp/view_department/';</script>")
    
def edit_department(request, department_id):
    try:
        department = tbl_department.objects.get(department_id=department_id)
    except tbl_department.DoesNotExist:
        return HttpResponse("<script>alert('Department not found');window.location='/adminapp/view_department/';</script>")

    if request.method == "POST":
        new_name = request.POST.get("department_name")

        if tbl_department.objects.filter(department_name=new_name).exclude(department_id=department_id).exists():
            return HttpResponse(
                "<script>alert('Department name already exists');window.location='/adminapp/department/edit/{}/';</script>".format(
                    department_id
                )
            )

        department.department_name = new_name
        department.save()

        return HttpResponse(
            "<script>alert('Department updated successfully');window.location='/adminapp/view_department/';</script>"
        )

    return render(request, "Admin/edit_department.html", {"department": department})

def classtype(request):
    return render(request, "Admin/classtype.html")

def classtype_insert(request):
    if request.method == "POST":
        name = request.POST.get("classtype_name")

        if tbl_classtype.objects.filter(classtype_name=name).exists():
            return HttpResponse("<script>alert('Class Type already exists');window.location='/adminapp/classtype/';</script>")

        tbl_classtype.objects.create(classtype_name=name)
        return HttpResponse("<script>alert('Class Type added');window.location='/adminapp/classtype/';</script>")

    return HttpResponse("<script>alert('Invalid Request');window.location='/adminapp/classtype/';</script>")


def view_classtype(request):
    data = tbl_classtype.objects.all().order_by('classtype_name')
    return render(request, "Admin/view_classtype.html", {"classtypes": data})

def delete_classtype(request, classtype_id):
    try:
        classtype = tbl_classtype.objects.get(classtype_id=classtype_id)
        classtype.delete()
        return HttpResponse("<script>alert('Class Type deleted successfully');window.location='/adminapp/view_classtype/';</script>")
    except tbl_classtype.DoesNotExist:
        return HttpResponse("<script>alert('Class Type not found');window.location='/adminapp/view_classtype/';</script>")

def edit_classtype(request, classtype_id):
    try:
        classtype = tbl_classtype.objects.get(classtype_id=classtype_id)
    except tbl_classtype.DoesNotExist:
        return HttpResponse("<script>alert('Class Type not found');window.location='/adminapp/view_classtype/';</script>")

    if request.method == "POST":
        new_name = request.POST.get("classtype_name")

        if tbl_classtype.objects.filter(classtype_name=new_name).exclude(classtype_id=classtype_id).exists():
            return HttpResponse("<script>alert('Class Type name already exists');window.location='/adminapp/edit_classtype/{}';</script>".format(classtype_id))

        classtype.classtype_name = new_name
        classtype.save()
        return HttpResponse("<script>alert('Class Type updated successfully');window.location='/adminapp/view_classtype/';</script>")

    return render(request, "Admin/edit_classtype.html", {"classtype": classtype})

def location(request):
    districts = tbl_district.objects.all()
    return render(request, "Admin/location.html", {'districts': districts})


def location_insert(request):
    if request.method == "POST":

        district_id = request.POST.get("ddldistrict")
        location_name = request.POST.get("locname")

        district_obj = tbl_district.objects.get(district_id=district_id)

        if tbl_location.objects.filter(location_name=location_name, district_id=district_id).exists():
            return HttpResponse("<script>alert('Location already exists');window.location='/adminapp/location/';</script>")

        obj = tbl_location(location_name=location_name, district_id=district_obj)
        obj.save()

        return HttpResponse("<script>alert('Location added');window.location='/adminapp/location/';</script>")

    return HttpResponse("<script>alert('Invalid Request');window.location='/adminapp/location/';</script>")


def view_location(request):
  district=tbl_district.objects.all()
  return render(request, "Admin/view_location.html", {"dis": district}) 

def filllocation(request):
    did=int(request.POST.get("did"))
    location = tbl_location.objects.filter(district_id=did).values()
    return JsonResponse(list(location), safe=False)

def delete_location(request, location_id):
    try:
        location = tbl_location.objects.get(location_id=location_id)
        location.delete()
        return HttpResponse("<script>alert('Location deleted successfully');window.location='/adminapp/view_location/';</script>")
    except tbl_location.DoesNotExist:
        return HttpResponse("<script>alert('Location not found');window.location='/adminapp/view_location/';</script>")
    
def edit_location(request, location_id):
    try:
        location = tbl_location.objects.get(location_id=location_id)
        districts = tbl_district.objects.all()
    except tbl_location.DoesNotExist:
        return HttpResponse("<script>alert('Location not found');window.location='/adminapp/view_location/';</script>")

    if request.method == "POST":
        new_name = request.POST.get("location_name")
        new_district_id = request.POST.get("district_id")

        district_obj = tbl_district.objects.get(district_id=new_district_id)

        if tbl_location.objects.filter(
            location_name=new_name,
            district_id=new_district_id
        ).exclude(location_id=location_id).exists():
            return HttpResponse(
                "<script>alert('Location already exists in this district');window.location='/adminapp/edit_location/{}/';</script>".format(location_id)
            )

        location.location_name = new_name
        location.district_id = district_obj
        location.save()

        return HttpResponse("<script>alert('Location updated successfully');window.location='/adminapp/view_location/';</script>")

    return render(request, "Admin/edit_location.html", {"location": location, "districts": districts})

def course(request):
    departments = tbl_department.objects.all()
    return render(request, "Admin/course.html", {"departments": departments})

def course_insert(request):
    if request.method == "POST":
        name = request.POST.get("course_name")
        dept_id = request.POST.get("department_id")

        dept_obj = tbl_department.objects.get(department_id=dept_id)

        if tbl_course.objects.filter(course_name=name, department_id=dept_obj).exists():
            return HttpResponse("<script>alert('Course already exists');window.location='/adminapp/course/';</script>")

        tbl_course.objects.create(course_name=name, department_id=dept_obj)
        return HttpResponse("<script>alert('Course added successfully');window.location='/adminapp/course/';</script>")

    return HttpResponse("<script>alert('Invalid Request');window.location='/adminapp/course/';</script>")
def view_course(request):
    data = tbl_department.objects.all()
    return render(request, "Admin/view_course.html", {"dept": data})

def fillcourse(request):
    did=int(request.POST.get("did"))
    course = tbl_course.objects.filter(department_id=did).values()
    return JsonResponse(list(course), safe=False)

def delete_course(request, course_id):
    
    try:
        course = tbl_course.objects.get(course_id=course_id)
        course.delete()
        return HttpResponse("<script>alert('Course deleted successfully');window.location='/adminapp/view_course/';</script>")
    except tbl_course.DoesNotExist:
        return HttpResponse("<script>alert('Course not found');window.location='/adminapp/view_course/';</script>")
    
def edit_course(request, course_id):
    try:
        course = tbl_course.objects.get(course_id=course_id)
        departments = tbl_department.objects.all()
    except tbl_course.DoesNotExist:
        return HttpResponse("<script>alert('Course not found');window.location='/adminapp/view_course/';</script>")

    if request.method == "POST":
        new_name = request.POST.get("course_name")
        new_department_id = request.POST.get("department_id")

        dept_obj = tbl_department.objects.get(department_id=new_department_id)

        if tbl_course.objects.filter(
            course_name=new_name,
            department_id=dept_obj
        ).exclude(course_id=course_id).exists():
            return HttpResponse(
                "<script>alert('Course already exists in this department');window.location='/adminapp/edit_course/{}/';</script>".format(course_id)
            )

        course.course_name = new_name
        course.department_id = dept_obj
        course.save()

        return HttpResponse("<script>alert('Course updated successfully');window.location='/adminapp/view_course/';</script>")

    return render(request, "Admin/edit_course.html", {"course": course, "departments": departments})

def trainingclass(request):
    batches = tbl_batch.objects.all()
    classtypes = tbl_classtype.objects.all()
      
    courses = tbl_course.objects.all()
    districts = tbl_district.objects.all()
    departments = tbl_department.objects.all()


    
    return render(request, "Admin/trainingclass.html", {
        "batches": batches,
        "classtypes": classtypes,
       
        "courses": courses,
        "districts": districts,
        "departments": departments,

    })

def training_class_insert(request):
    if request.method == "POST":
        tname = request.POST.get("trainingclass_name")
        stime = request.POST.get("start_time")
        tdate = request.POST.get("date")
        course_id = request.POST.get("course_id")
        classtype_id = request.POST.get("classtype_id")
        batch_id = request.POST.get("batch_id")
        
        desc = request.POST.get("description")

        course_obj = tbl_course.objects.get(course_id=course_id)
        classtype_obj = tbl_classtype.objects.get(classtype_id=classtype_id)
        batch_obj = tbl_batch.objects.get(batch_id=batch_id)
        

        tbl_trainingclass.objects.create(
            trainingclass_name=tname,
            start_time=stime,
            date=tdate,
            course_id=course_obj,
            classtype_id=classtype_obj,
            batch_id=batch_obj,
        
            description=desc
        )

        return HttpResponse("<script>alert('Training Class added successfully');window.location='/adminapp/trainingclass/';</script>")

    return HttpResponse("<script>alert('Invalid Request');window.location='/adminapp/trainingclass/';</script>")

def view_trainingclass(request):
    departments = tbl_department.objects.all()
    courses = tbl_course.objects.all()
    return render(request, "Admin/view_trainingclass.html", {
        "departments": departments,
        "courses": courses
    })

def filltraining(request):
    cid=int(request.POST.get("cid"))
    trainingclasses = tbl_trainingclass.objects.filter(course_id=cid).values('trainingclass_id', 'trainingclass_name', 'start_time', 'date', 'description', 'batch_id__batch_year', 'classtype_id__classtype_name')
    return JsonResponse(list(trainingclasses), safe=False)

def delete_training(request, trainingclass_id):
    try:
        trainingclass = tbl_trainingclass.objects.get(trainingclass_id=trainingclass_id)
        trainingclass.delete()
        return HttpResponse("<script>alert('Training Class deleted successfully');window.location='/adminapp/view_trainingclass/';</script>")    
    except tbl_trainingclass.DoesNotExist:
        return HttpResponse("<script>alert('Training Class not found');window.location='/adminapp/view_trainingclass/';</script>")
    
def edit_training(request, trainingclass_id):
    try:
        trainingclass = tbl_trainingclass.objects.get(trainingclass_id=trainingclass_id)
        batches = tbl_batch.objects.all()
        classtypes = tbl_classtype.objects.all()
        courses = tbl_course.objects.all()
    except tbl_trainingclass.DoesNotExist:
        return HttpResponse("<script>alert('Training Class not found');window.location='/adminapp/view_trainingclass/';</script>")

    if request.method == "POST":
        new_name = request.POST.get("trainingclass_name")
        new_start_time = request.POST.get("start_time")
        new_date = request.POST.get("date")
        new_course_id = request.POST.get("course_id")
        new_classtype_id = request.POST.get("classtype_id")
        new_batch_id = request.POST.get("batch_id")
        new_description = request.POST.get("description")

        course_obj = tbl_course.objects.get(course_id=new_course_id)
        classtype_obj = tbl_classtype.objects.get(classtype_id=new_classtype_id)
        batch_obj = tbl_batch.objects.get(batch_id=new_batch_id)

        trainingclass.trainingclass_name = new_name
        trainingclass.start_time = new_start_time
        trainingclass.date = new_date
        trainingclass.course_id = course_obj
        trainingclass.classtype_id = classtype_obj
        trainingclass.batch_id = batch_obj
        trainingclass.description = new_description
        trainingclass.save()

        return HttpResponse("<script>alert('Training Class updated successfully');window.location='/adminapp/view_trainingclass/';</script>")

    return render(request, "Admin/edit_training.html", {
        "trainingclass": trainingclass,
        "batches": batches,
        "classtypes": classtypes,
        "courses": courses
    })

def view_company(request):
    companies = tbl_company.objects.filter(login_id__status='requested').order_by('company_name')
    return render(request, "Admin/view_company.html", {"companies": companies})

def accept_company(request,loginid):
    login = tbl_login.objects.get(login_id=loginid)
    login.status = 'confirmed'
    login.save()
    return admin_dashboard(request)

def reject_company(request,loginid):
    login = tbl_login.objects.get(login_id=loginid)
    login.status = 'rejected'
    login.save()
    return admin_dashboard(request)

def registered_companies(request):
    companies = tbl_company.objects.filter(login_id__status='confirmed').order_by('-reg_date')
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date and to_date:
        companies = companies.filter(reg_date__range=[from_date, to_date])
    
    return render(request, 'Admin/registered_companies.html', {'companies': companies})

def companyexcel_export(request):
    companies = tbl_company.objects.filter(login_id__status='confirmed').order_by('-reg_date')
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date and to_date:
        companies = companies.filter(reg_date__range=[from_date, to_date])
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Companies"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Headers
    headers = ['Company Name', 'Contact Number', 'Email', 'Registration Date', 'Location']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Data
    for row_num, company in enumerate(companies, 2):
        ws.cell(row=row_num, column=1).value = company.company_name
        ws.cell(row=row_num, column=2).value = company.contact_number
        ws.cell(row=row_num, column=3).value = company.contact_email
        ws.cell(row=row_num, column=4).value = company.reg_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=5).value = company.location_id.location_name
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registered_companies.xlsx'
    
    wb.save(response)
    return response

def placed_students_barchart(request):
    """Simple bar chart showing placed students by course"""
    try:
        # Get all courses with placement data
        courses = tbl_course.objects.all().order_by('course_name')
        
        # Prepare data for chart
        chart_data = []
        for course in courses:
            # Count placed students for this course
            placed_count = tbl_student_schedule.objects.filter(
                student_id__course_id=course,
                status='selected'
            ).count()
            
            # Count total students in this course
            total_students = tbl_student.objects.filter(
                course_id=course,
                login_id__status='confirmed'
            ).count()
            
            # Calculate placement percentage
            placement_percent = 0
            if total_students > 0:
                placement_percent = round((placed_count / total_students) * 100, 2)
            
            chart_data.append({
                'course_name': course.course_name,
                'placed': placed_count,
                'total': total_students,
                'percentage': placement_percent
            })
        
        # Sort by placed count (highest first)
        chart_data.sort(key=lambda x: x['placed'], reverse=True)
        
        # Prepare data for Chart.js
        course_names = [item['course_name'] for item in chart_data]
        placed_counts = [item['placed'] for item in chart_data]
        percentages = [item['percentage'] for item in chart_data]
        
        context = {
            'chart_data': chart_data,
            'course_names_json': json.dumps(course_names),
            'placed_counts_json': json.dumps(placed_counts),
            'percentages_json': json.dumps(percentages),
            'total_placed': sum(item['placed'] for item in chart_data),
            'total_students': sum(item['total'] for item in chart_data),
        }
        
        return render(request, 'Admin/placed_students_barchart.html', context)
        
    except Exception as e:
        return HttpResponse(f"""
            <script>
                alert('Error: {str(e)}');
                window.location='/adminapp/admin_dashboard/';
            </script>
        """)
    # Get all courses
    courses = tbl_course.objects.all()
    
    course_data = []
    for course in courses:
        # Count students with status='selected'
        placed_count = tbl_student_schedule.objects.filter(
            student_id__course_id=course,
            status='selected'  # Make sure this is the correct status value
        ).count()
        
        # Also get total students in this course for reference
        total_students = tbl_student.objects.filter(
            course_id=course
        ).count()
        
        course_data.append({
            'name': course.course_name,
            'placed': placed_count,
            'total': total_students
        })
    
    # Sort by placed count (descending)
    course_data.sort(key=lambda x: x['placed'], reverse=True)
    
    # Calculate totals
    total_placed = sum(course['placed'] for course in course_data)
    total_courses = len(course_data)
    
    context = {
        'courses': course_data,
        'total_placed': total_placed,
        'total_courses': total_courses,
        'has_placements': total_placed > 0
    }
    
    return render(request, 'Admin/placed_students_chart.html', context)
    courses = tbl_course.objects.all()
    course_names = []
    placed_counts = []
    
    for course in courses:
        placed_count = tbl_student_schedule.objects.filter(
            student_id__course_id=course,
            status='selected'
        ).count()
        course_names.append(course.course_name)
        placed_counts.append(placed_count)
    
    context = {
        'course_names': json.dumps(course_names),
        'placed_counts': json.dumps(placed_counts),
    }
    return render(request, 'Admin/placed_students_chart.html', context)
      
    

def view_student(request):
    """View all students with their details"""
    status = request.GET.get('status', 'requested')
    
    students = tbl_student.objects.select_related(
        'login_id', 
        'course_id', 
        'batch_id'
    ).order_by('student_name')
    
    if status == 'confirmed':
        students = students.filter(login_id__status='confirmed')
    elif status == 'rejected':
        students = students.filter(login_id__status='rejected')
    else:
        students = students.filter(login_id__status='requested')
    
    total_students = tbl_student.objects.count()
    confirmed_count = tbl_student.objects.filter(login_id__status='confirmed').count()
    requested_count = tbl_student.objects.filter(login_id__status='requested').count()
    rejected_count = tbl_student.objects.filter(login_id__status='rejected').count()
    
    context = {
        'students': students,
        'status_filter': status,
        'total_students': total_students,
        'confirmed_count': confirmed_count,
        'requested_count': requested_count,
        'rejected_count': rejected_count,
    }
    return render(request, "Admin/view_student.html", context)


def accept_student(request, loginid):
    """Accept a student registration"""
    try:
        login = tbl_login.objects.get(login_id=loginid)
        login.status = 'confirmed'
        login.save()
        
        return HttpResponse("<script>alert('Student accepted successfully');window.location='/adminapp/view_student/';</script>")
    except tbl_login.DoesNotExist:
        return HttpResponse("<script>alert('Student not found');window.location='/adminapp/view_student/';</script>")


def reject_student(request, loginid):
    """Reject a student registration"""
    try:
        login = tbl_login.objects.get(login_id=loginid)
        login.status = 'rejected'
        login.save()
        
        return HttpResponse("<script>alert('Student rejected successfully');window.location='/adminapp/view_student/';</script>")
    except tbl_login.DoesNotExist:
        return HttpResponse("<script>alert('Student not found');window.location='/adminapp/view_student/';</script>")



def view_job_posts(request):
    """View all job posts with filters"""
    try:
        
        jobposts_list = tbl_jobpost.objects.all().order_by('-post_date')
        
        print(f"DEBUG: Found {jobposts_list.count()} jobs in database")
        
        if jobposts_list.exists():
            for job in jobposts_list:
                print(f"Job ID: {job.jobpost_id}, Position: {job.position}")
        
        status_filter = request.GET.get('status', '')
        if status_filter:
            jobposts_list = jobposts_list.filter(status=status_filter)
        
        search_query = request.GET.get('search', '')
        if search_query:
            jobposts_list = jobposts_list.filter(position__icontains=search_query)
        
        paginator = Paginator(jobposts_list, 10)
        page_number = request.GET.get('page')
        jobposts = paginator.get_page(page_number)
        
        companies = tbl_company.objects.filter(login_id__status='confirmed')
        
        today = date.today()
        print(f"DEBUG: Today is {today}")
        
        context = {
            'jobposts': jobposts,
            'status_filter': status_filter,
            'company_filter': '',
            'search_query': search_query,
            'companies': companies,
            'today': today,
        }
        return render(request, "Admin/view_job_posts.html", context)
        
    except Exception as e:
        print(f"ERROR in view_job_posts: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return HttpResponse(f"""
            <h1>Error loading job posts</h1>
            <p>Error: {str(e)}</p>
            <p><a href="/adminapp/admin_dashboard/">Back to Dashboard</a></p>
        """)

def job_details(request, jobpost_id):
    """View detailed information about a specific job post"""
    try:
        jobpost = tbl_jobpost.objects.select_related('company_id').get(jobpost_id=jobpost_id)
        
        company = jobpost.company_id
        
        total_students = tbl_student.objects.filter(login_id__status='confirmed').count()
        
        today = date.today()
        is_expired = jobpost.application_end_date < today
        
        context = {
            'jobpost': jobpost,
            'company': company,
            'total_students': total_students,
            'today': today,
            'is_expired': is_expired,
        }
        return render(request, "Admin/job_details.html", context)
        
    except tbl_jobpost.DoesNotExist:
        return HttpResponse(
            "<script>alert('Job post not found');window.location='/adminapp/job_posts/';</script>"
        )
def view_students_for_job(request, jobpost_id):
    """View students who meet the job requirements"""
    try:
        jobpost = tbl_jobpost.objects.get(jobpost_id=jobpost_id)
        
        print(f"DEBUG: Found jobpost - ID: {jobpost.jobpost_id}, Position: {jobpost.position}, Cutoff: {jobpost.cutoff_mark}")
        
        department = tbl_department.objects.all()
        batch = tbl_batch.objects.all()
        courses = tbl_course.objects.all()
        
        selected_department = request.GET.get('department', '')
        selected_batch = request.GET.get('batch', '')
        selected_course = request.GET.get('course', '')
        
        students = tbl_student.objects.filter(
            login_id__status='confirmed'
        ).select_related('course_id', 'batch_id', 'course_id__department_id')
        
        print(f"DEBUG: Total confirmed students: {students.count()}")
        
        if selected_department:
            students = students.filter(course_id__department_id=selected_department)
            print(f"DEBUG: After department filter: {students.count()}")
        
        if selected_batch:
            students = students.filter(batch_id=selected_batch)
            print(f"DEBUG: After batch filter: {students.count()}")
        
        if selected_course:
            students = students.filter(course_id=selected_course)
            print(f"DEBUG: After course filter: {students.count()}")
        
        meets_cutoff_count = 0
        below_cutoff_count = 0
        
        students_with_eligibility = []
        
        for student in students:
            student_score = student.percentage if student.percentage else 0
            meets_cutoff = student_score >= jobpost.cutoff_mark
            
            student_data = {
                'student': student,
                'meets_cutoff': meets_cutoff,
                'percentage': student_score,
            }
            students_with_eligibility.append(student_data)
            
            if meets_cutoff:
                meets_cutoff_count += 1
            else:
                below_cutoff_count += 1
            
            if len(students_with_eligibility) <= 3:
                print(f"DEBUG: Student {student.student_name} - Score: {student_score}, Cutoff: {jobpost.cutoff_mark}, Meets: {meets_cutoff}")
        
        print(f"DEBUG: Meets cutoff: {meets_cutoff_count}, Below cutoff: {below_cutoff_count}")
        
        context = {
            'jobpost': jobpost,
            'company': jobpost.company_id,
            'students_data': students_with_eligibility,
            'department': department,
            'batch': batch,
            'courses': courses,
            'total_students': len(students_with_eligibility),
            'meets_cutoff_count': meets_cutoff_count,
            'below_cutoff_count': below_cutoff_count,
            'selected_department': selected_department,
            'selected_batch': selected_batch,
            'selected_course': selected_course,
        }
        return render(request, "Admin/view_students_for_job.html", context)
        
    except tbl_jobpost.DoesNotExist:
        print(f"DEBUG: Job post not found - ID: {jobpost_id}")
        return HttpResponse(
            "<script>alert('Job post not found');window.location='/adminapp/job_posts/';</script>"
        )
    except Exception as e:
        print(f"DEBUG: Error in view_students_for_job: {str(e)}")
        return HttpResponse(
            f"<script>alert('Error: {str(e)}');window.location='/adminapp/job_posts/';</script>"
        )
def get_courses_by_department(request):
    """Get courses by department (AJAX)"""
    if request.method == "POST":
        department_id = request.POST.get('did')
        if department_id:
            courses = tbl_course.objects.filter(department_id=department_id).values('course_id', 'course_name')
            return JsonResponse(list(courses), safe=False)
    return JsonResponse([], safe=False)

def get_all_courses(request):
    """Get all courses (AJAX)"""
    if request.method == "POST":
        courses = tbl_course.objects.all().values('course_id', 'course_name')
        return JsonResponse(list(courses), safe=False)
    return JsonResponse([], safe=False)

def request_company(request):
    if request.method == "POST":
        try:
            jobpost_id = request.POST.get('jobpost_id')
            batch_id = request.POST.get('batch_id')
            student_count = int(request.POST.get('student_count', 0))

            if not jobpost_id or not batch_id or not student_count:
                return HttpResponse(
                    "<script>alert('Missing required fields');window.location='/adminapp/job_posts/';</script>"
                )

            jobpost = tbl_jobpost.objects.get(jobpost_id=jobpost_id)
            batch = tbl_batch.objects.get(batch_id=batch_id)

            # ✅ Correct validation call
            eligible_count = validate_request_creation(jobpost, batch, student_count)

            # ✅ Save batch in request
            request_obj = tbl_requests.objects.create(
                jobpost_id=jobpost,
                batch_id=batch,
                student_count=student_count,
                status=REQUEST_STATUS['PENDING']
            )

            return HttpResponse(
                f"<script>alert('Request #{request_obj.request_id} created successfully');"
                "window.location='/adminapp/job_posts/';</script>"
            )

        except Exception as e:
            return HttpResponse(
                f"<script>alert('Error: {str(e)}');"
                "window.location='/adminapp/job_posts/';</script>"
            )

    return HttpResponse(
        "<script>alert('Invalid request');window.location='/adminapp/job_posts/';</script>"
    )


def request_company_approval(request, jobpost_id):
    """Send selected students to company"""
    if request.method == "POST":
        try:
            jobpost = tbl_jobpost.objects.get(jobpost_id=jobpost_id)
            selected_students = request.POST.get('selected_students', '')
            student_count = request.POST.get('student_count', 0)
            message = request.POST.get('message', '')
            
            print(f"DEBUG: Sending {student_count} students to company")
            print(f"DEBUG: Selected students IDs: {selected_students}")
            print(f"DEBUG: Message: {message}")
            
           
            
            return HttpResponse(
                f"""
                <script>
                    alert('Successfully sent {student_count} student(s) to {jobpost.company_id.company_name} for approval.');
                    window.location='/adminapp/job_posts/';
                </script>
                """
            )
            
        except Exception as e:
            print(f"DEBUG: Error in request_company_approval: {str(e)}")
            return HttpResponse(
                f"""
                <script>
                    alert('Error: {str(e)}');
                    window.location='/adminapp/view_students_for_job/{jobpost_id}/';
                </script>
                """
            )
    
    return HttpResponse(
        "<script>alert('Invalid request');window.location='/adminapp/job_posts/';</script>"
    )



def view_company_accepted_students(request):
    try:
        today = date.today()

        all_requests = tbl_requests.objects.select_related(
            'jobpost_id',
            'jobpost_id__company_id'
        ).order_by('-request_date')

        companies = tbl_company.objects.filter(
            login_id__status='confirmed'
        ).order_by('company_name')

        request_data = []

        for req in all_requests:
            cutoff = req.jobpost_id.cutoff_mark or 0

            # GLOBAL eligible students
            eligible_students = tbl_student.objects.filter(
                login_id__status='confirmed',
                percentage__gte=cutoff
            )

            request_data.append({
                'request': req,
                'company': req.jobpost_id.company_id,
                'jobpost': req.jobpost_id,
                'eligible_students': eligible_students,
                'eligible_count': eligible_students.count(),
                'requested_count': req.student_count or 0,
            })

        status_counts = {
            'approved': tbl_requests.objects.filter(status='approved').count(),
            'pending': tbl_requests.objects.filter(status='pending').count(),
            'rejected': tbl_requests.objects.filter(status='rejected').count(),
        }

        context = {
            'request_data': request_data,
            'companies': companies,
            'total_requests': len(request_data),
            'status_counts': status_counts,
            'today': today,
        }

        return render(request, "Admin/company_accepted_students.html", context)

    except Exception as e:
        return HttpResponse(
            f"<script>alert('Error: {str(e)}');"
            "window.location='/adminapp/admin_dashboard/';</script>"
        )

def interview_schedule(request, id):
    schedule = tbl_requests.objects.get(request_id=id)
    
    students = tbl_student.objects.filter(
        batch_id=schedule.batch_id, 
        course_id=schedule.course_id,
        login_id__status='confirmed'
    )
    
    
    eligible_students = []
    for student in students:
        student_score = student.percentage if student.percentage else 0
        cutoff_mark = schedule.jobpost_id.cutoff_mark if schedule.jobpost_id.cutoff_mark else 0
        
        
        if student_score >= cutoff_mark:
            eligible_students.append(student)
    
    return render(request, "admin/interview_schedule.html", {
        'student': schedule, 
        'students': eligible_students
    })

def assign_students(request):
    if request.method == "POST":
        request_id = request.POST.get('request_id')
        selected_students = request.POST.getlist('selected_students')

        try:
            req = tbl_requests.objects.get(request_id=request_id)

            already_assigned = tbl_student_schedule.objects.filter(
                request_id=req
            ).exists()

            if already_assigned:
                return HttpResponse(
                    "<script>alert('Students already assigned for this request.');"
                    "window.location='/adminapp/company_accepted_students/';</script>"
                )

            for student_id in selected_students:
                student = tbl_student.objects.get(student_id=student_id)

                tbl_student_schedule.objects.create(
                    student_id=student,
                    request_id=req,
                    status=STUDENT_SCHEDULE_STATUS['ASSIGNED']
                )

            req.status = 'students_assigned'
            req.save()

            return HttpResponse(
                "<script>alert('Students assigned successfully');"
                "window.location='/adminapp/company_accepted_students/';</script>"
            )

        except Exception as e:
            return HttpResponse(
                f"<script>alert('Error: {str(e)}');"
                "window.location='/adminapp/company_accepted_students/';</script>"
            )

    return HttpResponse(
        "<script>alert('Invalid request');"
        "window.location='/adminapp/company_accepted_students/';</script>"
    )


def change_request_status(request, request_id, new_status):
    try:
        req = tbl_requests.objects.get(request_id=request_id)
        req.status = new_status
        req.save()
        return HttpResponse(f"<script>alert('Status updated');window.location='/adminapp/company_accepted_students/';</script>")
    except:
        return HttpResponse(f"<script>alert('Error');window.location='/adminapp/company_accepted_students/';</script>")

def request_again(request, request_id):
    return change_request_status(request, request_id, 'pending')

def reject_accepted_request(request, request_id):
    return change_request_status(request, request_id, 'rejected')

def approve_request(request, request_id):
    return change_request_status(request, request_id, 'approved')




def view_students_scheduled_by_company(request):
    try:
        student_schedules = tbl_student_schedule.objects.select_related(
            'student_id',
            'student_id__batch_id',
            'student_id__course_id',
            'request_id',
            'request_id__jobpost_id',
            'request_id__jobpost_id__company_id',
            'schedule_id'
        ).all().order_by('-assigned_date')
        
        company_filter = request.GET.get('company', '')
        status_filter = request.GET.get('status', '')
        batch_filter = request.GET.get('batch', '')
        date_filter = request.GET.get('date', '')
        
        if company_filter:
            student_schedules = student_schedules.filter(
                request_id__jobpost_id__company_id=company_filter
            )
        
        if status_filter:
            student_schedules = student_schedules.filter(status=status_filter)
        
        if batch_filter:
            student_schedules = student_schedules.filter(
                student_id__batch_id=batch_filter
            )
        
        if date_filter:
            student_schedules = student_schedules.filter(
                assigned_date=date_filter
            )
        
        companies = tbl_company.objects.filter(login_id__status='confirmed')
        batches = tbl_batch.objects.all()
        
        total_scheduled = student_schedules.count()
        assigned_count = student_schedules.filter(status='assigned').count()
        interviewed_count = student_schedules.filter(status='interviewed').count()
        selected_count = student_schedules.filter(status='selected').count()
        rejected_count = student_schedules.filter(status='rejected').count()
        
        context = {
            'student_schedules': student_schedules,
            'companies': companies,
            'batches': batches,
            'company_filter': company_filter,
            'status_filter': status_filter,
            'batch_filter': batch_filter,
            'date_filter': date_filter,
            'total_scheduled': total_scheduled,
            'assigned_count': assigned_count,
            'interviewed_count': interviewed_count,
            'selected_count': selected_count,
            'rejected_count': rejected_count,
        }
        
        return render(request, "Admin/view_students_scheduled.html", context)
        
    except Exception as e:
        return HttpResponse(f"""
            <script>
                alert('Error: {str(e)}');
                window.location='/adminapp/dashboard/';
            </script>
        """)
    

def student_distribution_piechart(request):
    """Generate pie chart showing student distribution by course"""
    try:
        # Get course-wise student count
        queryset = tbl_student.objects.filter(
            login_id__status='confirmed'
        ).values(
            'course_id__course_name'
        ).annotate(
            student_count=Count('student_id')
        ).order_by('course_id__course_name')
        
        # Prepare data
        course_data = []
        labels = []
        data = []
        
        for item in queryset:
            if item['course_id__course_name']:
                course_name = item['course_id__course_name']
                count = item['student_count']
                
                course_data.append((course_name, count))
                labels.append(course_name)
                data.append(count)
        
        # Calculate totals
        total_students = sum(data) if data else 0
        total_courses = len(labels)
        all_courses_count = tbl_course.objects.count()
        
        context = {
            'course_data': course_data,
            'labels': labels,
            'data': data,
            'labels_json': json.dumps(labels),
            'data_json': json.dumps(data),
            'total_students': total_students,
            'total_courses': total_courses,
            'all_courses_count': all_courses_count,
        }
        
        return render(request, 'Admin/student_distribution_piechart.html', context)
        
    except Exception as e:
        print(f"Error in student_distribution_piechart: {str(e)}")
        # Return simple error context
        context = {
            'course_data': [],
            'labels': [],
            'data': [],
            'labels_json': '[]',
            'data_json': '[]',
            'total_students': 0,
            'total_courses': 0,
            'all_courses_count': 0,
        }
        return render(request, 'Admin/student_distribution_piechart.html', context)